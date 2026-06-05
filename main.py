import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.core.window import Window
from kivy.utils import get_color_from_hex

# =================================================================
# ⚙️ دالة تصدير الإكسيل (نفس الدالة الاحترافية للكمبيوتر)
# =================================================================
def generate_formatted_excel(grid_data):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "مخطط خلايا المخزن"
        
        ws.sheet_view.rightToLeft = True
        ws.merge_cells("A1:J1")
        ws["A1"] = "مخطط التوزيع البصري لخلايا مخزن دمياط (موبايل)"
        ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1A365D")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        color_map = {
            "إصابة حية": {"fill": "FFC7CE", "font": "9C0006"},
            "إصابة ميتة": {"fill": "FFEB9C", "font": "9C6500"},
            "تبخير": {"fill": "F8CBAD", "font": "C65911"},
            "سليم": {"fill": "C6EFCE", "font": "006100"},
            "empty": {"fill": "F2F2F2", "font": "7F7F7F"}
        }
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for r in range(4):
            ws.row_dimensions[r+3].height = 55
            for c in range(10):
                cell = ws.cell(row=r+3, column=c+1)
                ws.column_dimensions[get_column_letter(c+1)].width = 24
                
                cell_info = grid_data[r][c]
                product = cell_info.get("product", "").strip()
                status = cell_info.get("status", "").strip()

                if product:
                    cell.value = f"{product}\n({status})"
                    style = color_map["empty"]
                    for key, val in color_map.items():
                        if key in status: style = val; break
                    cell.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
                    cell.font = Font(name="Arial", size=10, bold=True, color=style["font"])
                else:
                    cell.value = f"خلية {r+1}-{c+1}\n(فارغة)"
                    cell.fill = PatternFill(start_color=color_map["empty"]["fill"], end_color=color_map["empty"]["fill"], fill_type="solid")
                    cell.font = Font(name="Arial", size=9, italic=True, color=color_map["empty"]["font"])
                
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

        filename = "تقرير_تنسيق_المخزن_موبايل.xlsx"
        wb.save(filename)
        return True, filename
    except Exception as e:
        return False, str(e)


# =================================================================
# 📱 كلاس واجهة الموبايل (Android UI)
# =================================================================
class DamyattaAndroidApp(App):
    def build(self):
        Window.clearcolor = get_color_from_hex("#f8f9fa")
        
        self.grid_data = [[{"product": "", "status": ""} for _ in range(10)] for _ in range(4)]
        self.ui_colors = {
            "إصابة حية": {"bg": "#FFC7CE", "fg": "#9C0006"},
            "إصابة ميتة": {"bg": "#FFEB9C", "fg": "#9C6500"},
            "تبخير": {"bg": "#F8CBAD", "fg": "#C65911"},
            "سليم": {"bg": "#C6EFCE", "fg": "#006100"},
            "empty": {"bg": "#edf2f7", "fg": "#555555"}
        }

        # التخطيط الرئيسي للشاشة
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        # 1. عنوان التطبيق
        title = Label(text="Damyatta Warehouse - Mobile", size_hint_y=0.1, color=get_color_from_hex("#1a365d"), bold=True, font_size=20)
        main_layout.add_widget(title)

        # 2. شبكة الخلايا (مع خاصية السحب باللمس)
        scroll = ScrollView(size_hint_y=0.5, do_scroll_x=True, do_scroll_y=True)
        grid = GridLayout(cols=10, spacing=5, size_hint_x=None, size_hint_y=None)
        grid.bind(minimum_width=grid.setter('width'), minimum_height=grid.setter('height'))

        self.btn_grid = []
        for r in range(4):
            row_btns = []
            for c in range(10):
                btn = Button(
                    text=f"Cell {r+1}-{c+1}\n(Empty)", 
                    size_hint=(None, None), size=(120, 100),
                    background_normal="",
                    background_color=get_color_from_hex(self.ui_colors["empty"]["bg"]),
                    color=get_color_from_hex(self.ui_colors["empty"]["fg"]),
                    bold=True
                )
                btn.bind(on_press=lambda instance, row=r+1, col=c+1: self.on_cell_click(row, col))
                grid.add_widget(btn)
                row_btns.append(btn)
            self.btn_grid.append(row_btns)
            
        scroll.add_widget(grid)
        main_layout.add_widget(scroll)

        # 3. لوحة الإدخال
        input_layout = GridLayout(cols=2, size_hint_y=0.3, spacing=10)
        
        input_layout.add_widget(Label(text="Row (1-4):", color=(0,0,0,1), bold=True))
        self.entry_row = TextInput(multiline=False, input_filter='int', font_size=18, halign="center")
        input_layout.add_widget(self.entry_row)

        input_layout.add_widget(Label(text="Column (1-10):", color=(0,0,0,1), bold=True))
        self.entry_col = TextInput(multiline=False, input_filter='int', font_size=18, halign="center")
        input_layout.add_widget(self.entry_col)

        products = ["عدس مجروش استرالي", "عدس جبه استرالي", "فول مدشوش", "حصى استرالي"]
        self.combo_product = Spinner(text="Product Name", values=products, background_normal="", background_color=(0.8,0.8,0.8,1), color=(0,0,0,1))
        input_layout.add_widget(self.combo_product)

        statuses = ["سليم", "إصابة حية", "إصابة ميتة", "تبخير"]
        self.combo_status = Spinner(text="Status", values=statuses, background_normal="", background_color=(0.8,0.8,0.8,1), color=(0,0,0,1))
        input_layout.add_widget(self.combo_status)

        main_layout.add_widget(input_layout)

        # 4. أزرار الحفظ والتصدير
        btn_layout = BoxLayout(orientation='horizontal', size_hint_y=0.1, spacing=10)
        
        btn_save = Button(text="Save Cell", background_normal="", background_color=get_color_from_hex("#007bff"), bold=True)
        btn_save.bind(on_press=self.save_data)
        btn_layout.add_widget(btn_save)

        btn_excel = Button(text="Export Excel", background_normal="", background_color=get_color_from_hex("#217346"), bold=True)
        btn_excel.bind(on_press=self.export_excel)
        btn_layout.add_widget(btn_excel)

        main_layout.add_widget(btn_layout)
        
        self.load_data()
        return main_layout

    def on_cell_click(self, r, c):
        self.entry_row.text = str(r)
        self.entry_col.text = str(c)
        current_cell = self.grid_data[r-1][c-1]
        if current_cell["product"]:
            self.combo_product.text = current_cell["product"]
            self.combo_status.text = current_cell["status"]
        else:
            self.combo_product.text = "Product Name"
            self.combo_status.text = "Status"

    def save_data(self, instance):
        try:
            row = int(self.entry_row.text)
            col = int(self.entry_col.text)
            product = self.combo_product.text
            status = self.combo_status.text
            
            if not (1 <= row <= 4) or not (1 <= col <= 10): return
            if product == "Product Name" or status == "Status": return

            style = self.ui_colors["empty"]
            for key, val in self.ui_colors.items():
                if key in status: style = val; break
            
            display_text = f"Cell {row}-{col}\n{product}\n({status})"
            self.btn_grid[row-1][col-1].text = display_text
            self.btn_grid[row-1][col-1].background_color = get_color_from_hex(style["bg"])
            self.btn_grid[row-1][col-1].color = get_color_from_hex(style["fg"])
            
            self.grid_data[row-1][col-1] = {"product": product, "status": status}
            
            # حفظ في الذاكرة الداخلية للموبايل
            with open("warehouse_log.txt", "a", encoding="utf-8") as f:
                f.write(f"{row}|{col}|{product}|{status}\n")
        except ValueError:
            pass

    def load_data(self):
        if os.path.exists("warehouse_log.txt"):
            with open("warehouse_log.txt", "r", encoding="utf-8") as f:
                for line in f:
                    if not line.strip(): continue
                    try:
                        parts = line.strip().split("|")
                        row, col, product, status = int(parts[0]), int(parts[1]), parts[2], parts[3]
                        self.grid_data[row-1][col-1] = {"product": product, "status": status}
                        
                        style = self.ui_colors["empty"]
                        for key, val in self.ui_colors.items():
                            if key in status: style = val; break
                        
                        self.btn_grid[row-1][col-1].text = f"Cell {row}-{col}\n{product}\n({status})"
                        self.btn_grid[row-1][col-1].background_color = get_color_from_hex(style["bg"])
                        self.btn_grid[row-1][col-1].color = get_color_from_hex(style["fg"])
                    except:
                        pass

    def export_excel(self, instance):
        success, res = generate_formatted_excel(self.grid_data)
        if success:
            print("Exported successfully to:", res)

if __name__ == '__main__':
    DamyattaAndroidApp().run()
